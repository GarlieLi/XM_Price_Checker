# Previous version, not using anymore.
# Just keeping it for reference.

from pathlib import Path
from datetime import datetime
import re
import time

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


# ============================================================
# 1. Open Excel workbook
# ============================================================

project_folder = Path(__file__).parent
excel_file = project_folder / "XM_Price_Checker_Python.xlsm"

workbook = load_workbook(
    excel_file,
    keep_vba=True
)

links_sheet = workbook["Links"]
rawdata_sheet = workbook["RawData"]
products_sheet = workbook["Products"]


# ============================================================
# 2. Helper functions
# ============================================================

def clean_price(price_text):
    """
    Convert price text such as:

        499
        499.00
        499,00
        499 zł
        Cena: 499,00 zł

    into a float.
    """

    if price_text is None:
        raise ValueError("Price text is empty.")

    text = str(price_text).strip()

    text = text.replace("Cena:", "")
    text = text.replace("zł", "")
    text = text.replace("PLN", "")

    text = re.sub(r"\s+", "", text)

    text = re.sub(r"[^0-9,.]", "", text)

    if not text:
        raise ValueError(
            f"Could not extract numeric price from: {price_text}"
        )

    if "," in text and "." not in text:
        text = text.replace(",", ".")

    elif "," in text and "." in text:
        text = text.replace(".", "")
        text = text.replace(",", ".")

    return float(text)


# ============================================================
# 3. Read product information
# ============================================================

def get_product(target_id):

    headers = []

    for cell in products_sheet[1]:

        if cell.value is None:
            headers.append("")

        else:
            headers.append(
                str(cell.value).strip()
            )

    header_map = {
        header.upper(): index
        for index, header in enumerate(headers)
        if header
    }

    product_model_col = header_map.get(
        "PRODUCT MODEL"
    )

    product_name_col = header_map.get(
        "PRODUCT NAME"
    )

    ram_col = header_map.get(
        "RAM"
    )

    storage_col = header_map.get(
        "STORAGE"
    )

    if product_model_col is None:
        raise ValueError(
            "Product Model column not found in Products."
        )

    if product_name_col is None:
        raise ValueError(
            "Product Name column not found in Products."
        )

    if ram_col is None:
        raise ValueError(
            "RAM column not found in Products."
        )

    if storage_col is None:
        raise ValueError(
            "Storage column not found in Products."
        )

    target_id_clean = str(
        target_id
    ).strip().upper()

    for row_number, row in enumerate(
        products_sheet.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):

        if not row:
            continue

        product_model = row[
            product_model_col
        ]

        ram = row[
            ram_col
        ]

        storage = row[
            storage_col
        ]

        if (
            product_model is None
            or ram is None
            or storage is None
        ):
            continue

        product_model = str(
            product_model
        ).strip()

        ram = str(
            ram
        ).strip()

        storage = str(
            storage
        ).strip()

        # Same logic as Excel:
        #
        # Product Model + RAM + Storage
        #
        calculated_target_id = (
            f"{product_model}-{ram}-{storage}"
        )

        if (
            calculated_target_id.strip().upper()
            == target_id_clean
        ):

            product_name = (
                str(
                    row[product_name_col]
                ).strip()
                if row[product_name_col] is not None
                else product_model
            )

            print(
                "Product found:",
                product_name,
                "| RAM:",
                ram,
                "| Storage:",
                storage
            )

            return {
                "target_id": target_id,
                "name": product_name,
                "ram": ram,
                "storage": storage
            }

    raise ValueError(
        f"TargetID not found in Products: {target_id}"
    )


# ============================================================
# 4. MEX scraper
# ============================================================

def get_mediaexpert_price(page):

    price_elements = page.locator(
        "span.whole"
    )

    if price_elements.count() == 0:
        raise ValueError(
            "Media Expert price element not found."
        )

    whole = (
        price_elements.first
        .inner_text()
        .strip()
    )

    decimal_elements = page.locator(
        "span.decimal"
    )

    if decimal_elements.count() > 0:

        decimal = (
            decimal_elements.first
            .inner_text()
            .strip()
        )

    else:

        decimal = "00"

    return float(
        f"{whole}.{decimal}"
    )


# ============================================================
# 5. XKOM scraper
# ============================================================

def get_xkom_price(page):

    price_element = page.locator(
        "span.parts__Price-sc-fd70cef5-1"
    ).first

    if price_element.count() == 0:
        raise ValueError(
            "X-kom price element not found."
        )

    whole = (
        price_element
        .inner_text()
        .strip()
    )

    return float(whole)


# ============================================================
# 6. ELECTRO scraper
# ============================================================

def get_electro_price(page):

    price_elements = page.locator(
        "span.whole"
    )

    if price_elements.count() == 0:
        raise ValueError(
            "Electro price element not found."
        )

    whole = (
        price_elements.first
        .inner_text()
        .strip()
    )

    decimal_elements = page.locator(
        "span.decimal"
    )

    if decimal_elements.count() > 0:

        decimal = (
            decimal_elements.first
            .inner_text()
            .strip()
        )

    else:

        decimal = "00"

    return float(
        f"{whole}.{decimal}"
    )


# ============================================================
# 7. KOMPUTRONIK scraper
# ============================================================

def get_ktr_price(page):

    price_element = page.locator(
        '[data-price-type="final"]'
    ).first

    if price_element.count() == 0:
        raise ValueError(
            "Komputronik price element not found."
        )

    price_text = (
        price_element
        .inner_text()
        .strip()
    )

    return clean_price(
        price_text
    )


# ============================================================
# 8. NEONET scraper
# ============================================================

def get_neonet_results(page, product):

    results = []

    print("Waiting for NEONET products...")

    product_selector = "h3[class*='parts__Title']"

    # --------------------------------------------------------
    # First, wait for either:
    #   1. Product cards
    #   2. Verification page
    # --------------------------------------------------------

    try:

        page.locator(
            product_selector
        ).first.wait_for(
            state="visible",
            timeout=15000
        )

    except Exception:

        page_text = ""

        try:
            page_text = page.locator("body").inner_text()
        except Exception:
            pass

        # ----------------------------------------------------
        # Detect NEONET human verification
        # ----------------------------------------------------

        verification_detected = (
            "Verifying you are human" in page_text
            or
            "security service to protect" in page_text
            or
            "you are not a bot" in page_text
        )

        if verification_detected:

            print()
            print("========================================")
            print("NEONET HUMAN VERIFICATION DETECTED")
            print("Please complete the verification")
            print("in the browser window.")
            print("The script will wait for you...")
            print("========================================")
            print()

            # ------------------------------------------------
            # Wait up to 120 seconds for product cards
            # to appear after manual verification.
            # ------------------------------------------------

            try:

                page.locator(
                    product_selector
                ).first.wait_for(
                    state="visible",
                    timeout=120000
                )

                print(
                    "NEONET verification completed."
                )

            except Exception:

                raise ValueError(
                    "NEONET verification was not completed "
                    "within 120 seconds."
                )

        else:

            raise ValueError(
                "NEONET product cards did not appear."
            )

    # --------------------------------------------------------
    # Product cards should now be available
    # --------------------------------------------------------

    titles = page.locator(
        product_selector
    )

    print(
        "NEONET total product titles:",
        titles.count()
    )

    # --------------------------------------------------------
    # Target RAM / Storage
    # --------------------------------------------------------

    target_ram = re.sub(
        r"\D",
        "",
        product["ram"]
    )

    target_storage = re.sub(
        r"\D",
        "",
        product["storage"]
    )

    # --------------------------------------------------------
    # Process products
    # --------------------------------------------------------

    for i in range(titles.count()):

        title_element = titles.nth(i)

        title = (
            title_element
            .inner_text()
            .strip()
        )

        print()
        print(
            "NEONET title:",
            title
        )

        if (
            product["name"].lower()
            not in title.lower()
        ):
            continue

        # Find product card
        card = title_element.locator(
            "xpath=ancestor::div[contains(@class, 'parts__ContentWrapper')]"
        ).first

        if card.count() == 0:
            continue

        card_text = card.inner_text()

        # RAM
        ram_match = re.search(
            r"Pamięć RAM:\s*(\d+)\s*GB",
            card_text,
            re.IGNORECASE
        )

        # Storage
        storage_match = re.search(
            r"Pamięć wbudowana:\s*(\d+)\s*GB",
            card_text,
            re.IGNORECASE
        )

        if not ram_match or not storage_match:
            continue

        ram = ram_match.group(1)
        storage = storage_match.group(1)

        if (
            ram != target_ram
            or storage != target_storage
        ):
            continue

        # Color
        color = ""

        color_match = re.search(
            r"\b("
            r"Black|Green|Blue|White|"
            r"Gray|Grey|Purple|Silver|"
            r"Gold|Pink|Red"
            r")\b",
            title,
            re.IGNORECASE
        )

        if color_match:
            color = color_match.group(1)

        # Price
        price_element = card.locator(
            'div[data-name="productPrice"] '
            'span[aria-label^="Cena:"]'
        ).first

        if price_element.count() == 0:
            continue

        price_text = (
            price_element
            .get_attribute("aria-label")
        )

        price = clean_price(
            price_text
        )

        print(
            f"Matched: {color} | "
            f"{ram} GB / {storage} GB | "
            f"{price} PLN"
        )

        results.append({
            "product_name": product["name"],
            "variant": color,
            "price": price
        })

    return results


# ============================================================
# 9. Read enabled websites from Links
# ============================================================

links_to_check = []

for row in links_sheet.iter_rows(
    values_only=True
):

    if not row:
        continue

    target_id = row[0]
    website = row[1]
    url = row[2]
    enabled = row[3]

    if (
        str(enabled)
        .strip()
        .upper() == "YES"

        and url

        and str(website)
        .strip()
        .upper()
        in [
            "MEX",
            "XKOM",
            "ELECTRO",
            "KTR",
            "NEONET"
        ]
    ):

        links_to_check.append({

            "target_id":
                target_id,

            "website":
                str(website)
                .strip()
                .upper(),

            "url":
                url
        })


print()
print(
    "Links to check:",
    len(links_to_check)
)


# ============================================================
# 10. Start browser
# ============================================================

with sync_playwright() as p:

    browser = p.chromium.launch(
        headless=False
    )

    page = browser.new_page()

    for item in links_to_check:

        target_id = item[
            "target_id"
        ]

        website = item[
            "website"
        ]

        url = item[
            "url"
        ]

        print()
        print(
            "========================================"
        )

        print(
            "TargetID:",
            target_id
        )

        print(
            "Website:",
            website
        )

        print(
            "URL:",
            url
        )

        # ----------------------------------------------------
        # Get product
        # ----------------------------------------------------

        try:

            product = get_product(
                target_id
            )

        except Exception as e:

            print(
                "ERROR:",
                e
            )

            continue

        run_time = datetime.now()

        # ====================================================
        # Open website
        # ====================================================

        try:

            print(
                "Opening page..."
            )

            page.goto(
                url,
                wait_until="domcontentloaded",
                timeout=60000
            )

            print(
                "Page loaded."
            )

            # Give dynamically rendered pages
            # some time to finish loading.

            if website == "NEONET":

                page.wait_for_timeout(
                    3000
                )

            # =================================================
            # NEONET
            # =================================================

            if website == "NEONET":
                time.sleep(5)

                results = get_neonet_results(
                    page,
                    product
                )

                if not results:

                    raise ValueError(
                        "No matching NEONET "
                        "product found."
                    )

                # ------------------------------------------------
                # IMPORTANT:
                #
                # Every color is written separately.
                #
                # We DO NOT select the lowest price here.
                # ------------------------------------------------

                for result in results:

                    rawdata_sheet.append([

                        run_time,

                        target_id,

                        website,

                        result[
                            "product_name"
                        ],

                        result[
                            "variant"
                        ],

                        url,

                        result[
                            "price"
                        ],

                        "PLN",

                        "OK",

                        ""
                    ])

                    print(
                        "RawData added:",
                        result
                    )

            # =================================================
            # MEX
            # =================================================

            elif website == "MEX":

                price = get_mediaexpert_price(
                    page
                )

                print(
                    "Price:",
                    price
                )

                rawdata_sheet.append([

                    run_time,
                    target_id,
                    website,
                    product["name"],
                    "",
                    url,
                    price,
                    "PLN",
                    "OK",
                    ""
                ])

            # =================================================
            # XKOM
            # =================================================

            elif website == "XKOM":

                price = get_xkom_price(
                    page
                )

                print(
                    "Price:",
                    price
                )

                rawdata_sheet.append([

                    run_time,
                    target_id,
                    website,
                    product["name"],
                    "",
                    url,
                    price,
                    "PLN",
                    "OK",
                    ""
                ])

            # =================================================
            # ELECTRO
            # =================================================

            elif website == "ELECTRO":

                price = get_electro_price(
                    page
                )

                print(
                    "Price:",
                    price
                )

                rawdata_sheet.append([

                    run_time,
                    target_id,
                    website,
                    product["name"],
                    "",
                    url,
                    price,
                    "PLN",
                    "OK",
                    ""
                ])

            # =================================================
            # KOMPUTRONIK
            # =================================================

            elif website == "KTR":

                price = get_ktr_price(
                    page
                )

                print(
                    "Price:",
                    price
                )

                rawdata_sheet.append([

                    run_time,
                    target_id,
                    website,
                    product["name"],
                    "",
                    url,
                    price,
                    "PLN",
                    "OK",
                    ""
                ])

            else:

                raise ValueError(
                    f"No scraper available for {website}"
                )

        except Exception as e:

            error_message = str(e)

            print(
                "ERROR:",
                error_message
            )

            # ------------------------------------------------
            # Record failed website
            # ------------------------------------------------

            rawdata_sheet.append([

                run_time,
                target_id,
                website,
                product["name"],
                "",
                url,
                None,
                "PLN",
                "ERROR",
                error_message
            ])


    # ========================================================
    # Close browser
    # ========================================================

    browser.close()


# ============================================================
# 11. Save workbook
# ============================================================

workbook.save(
    excel_file
)

print()
print(
    "========================================"
)

print(
    "Finished."
)

print(
    "RawData updated."
)

print(
    "Workbook saved."
)